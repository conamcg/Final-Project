import anthropic
import os
import csv
import io
import datetime
import streamlit as st

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Cell Therapy Ticket Triage",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:ital,wght@0,300;0,400;0,500;1,300&family=Syne:wght@400;600;700;800&display=swap');

:root {
    --bg:        #0d0f14;
    --surface:   #13161e;
    --border:    #1e2330;
    --accent:    #00e5a0;
    --danger:    #ff4d6d;
    --muted:     #8b92a8;
    --text:      #e8eaf0;
    --subtext:   #c8ccd8;
    --mono:      'DM Mono', monospace;
    --display:   'Syne', sans-serif;
}

html, body, [data-testid="stAppViewContainer"] {
    background: var(--bg) !important;
    color: var(--text) !important;
}

[data-testid="stAppViewContainer"] > .main {
    background: var(--bg) !important;
}

/* Hide default streamlit chrome */
#MainMenu, footer, header { visibility: hidden; }
[data-testid="stToolbar"] { display: none; }

/* Typography - targeted, not broad */
h1, h2, h3, h4 {
    font-family: var(--display) !important;
    letter-spacing: -0.02em;
}

/* Header banner */
.header-banner {
    background: linear-gradient(135deg, #0d0f14 0%, #13161e 50%, #0d1a14 100%);
    border-bottom: 1px solid var(--border);
    padding: 2rem 0 1.5rem 0;
    margin-bottom: 2rem;
}

.header-title {
    font-family: var(--display) !important;
    font-size: 2.2rem;
    font-weight: 800;
    color: var(--text);
    letter-spacing: -0.03em;
    margin: 0;
    line-height: 1;
}

.header-title span {
    color: var(--accent);
    font-family: var(--display) !important;
}

.header-sub {
    font-family: var(--mono) !important;
    font-size: 0.75rem;
    color: var(--subtext);
    margin-top: 0.4rem;
    letter-spacing: 0.08em;
    text-transform: uppercase;
}

/* Cards */
.card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 1.5rem;
    margin-bottom: 1rem;
}

/* Classification badges */
.badge-escalate {
    display: inline-block;
    background: rgba(255, 77, 109, 0.12);
    color: var(--danger);
    border: 1px solid rgba(255, 77, 109, 0.3);
    border-radius: 4px;
    padding: 0.2rem 0.7rem;
    font-family: var(--mono) !important;
    font-size: 0.75rem;
    font-weight: 500;
    letter-spacing: 0.1em;
    text-transform: uppercase;
}

.badge-delegate {
    display: inline-block;
    background: rgba(0, 229, 160, 0.08);
    color: var(--accent);
    border: 1px solid rgba(0, 229, 160, 0.25);
    border-radius: 4px;
    padding: 0.2rem 0.7rem;
    font-family: var(--mono) !important;
    font-size: 0.75rem;
    font-weight: 500;
    letter-spacing: 0.1em;
    text-transform: uppercase;
}

.badge-vip {
    display: inline-block;
    background: rgba(255, 200, 0, 0.1);
    color: #ffc800;
    border: 1px solid rgba(255, 200, 0, 0.25);
    border-radius: 4px;
    padding: 0.2rem 0.6rem;
    font-family: var(--mono) !important;
    font-size: 0.7rem;
    font-weight: 500;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    margin-left: 0.5rem;
}

/* Rubric grid */
.rubric-grid {
    display: grid;
    grid-template-columns: 1fr 1fr 1fr;
    gap: 0.5rem;
    margin: 1rem 0;
}

.rubric-item {
    background: #0d0f14;
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 0.6rem 0.8rem;
    display: flex;
    justify-content: space-between;
    align-items: center;
}

.rubric-label {
    font-family: var(--mono) !important;
    font-size: 0.7rem;
    color: var(--text);
}

.rubric-yes {
    font-family: var(--mono) !important;
    font-size: 0.7rem;
    color: var(--danger);
    font-weight: 500;
}

.rubric-no {
    font-family: var(--mono) !important;
    font-size: 0.7rem;
    color: var(--muted);
}

/* Disclaimer */
.disclaimer {
    background: rgba(255, 77, 109, 0.05);
    border: 1px solid rgba(255, 77, 109, 0.2);
    border-radius: 6px;
    padding: 0.75rem 1rem;
    font-family: var(--mono) !important;
    font-size: 0.72rem;
    color: #ff8fa3;
    line-height: 1.5;
    margin-bottom: 1.5rem;
}

/* ── Email draft text areas — white text ── */
.stTextArea textarea,
.stTextArea textarea:disabled,
.stTextArea [disabled] {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    opacity: 1 !important;
}
/* Email section captions (titles) — white */
.email-caption {
    font-family: 'DM Mono', monospace !important;
    font-size: 0.72rem !important;
    font-weight: 700 !important;
    letter-spacing: 0.08em !important;
    text-transform: uppercase !important;
    color: #ffffff !important;
    margin-bottom: 0.25rem !important;
    margin-top: 0.75rem !important;
}
    background: #0a0c11 !important;
    border: 1px solid var(--border) !important;
    border-radius: 6px !important;
    color: var(--text) !important;
    font-family: var(--mono) !important;
    font-size: 0.85rem !important;
}

.stTextArea textarea:focus {
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 2px rgba(0, 229, 160, 0.1) !important;
}

.stTextInput input {
    background: #0a0c11 !important;
    border: 1px solid var(--border) !important;
    color: var(--text) !important;
    font-family: var(--mono) !important;
    font-size: 0.85rem !important;
}

/* ── Unified button style ── */
/* Default + primary buttons: GREEN */
div.stButton > button,
div.stButton > button[kind="primary"],
div.stButton > button[kind="primaryFormSubmit"],
[data-testid="stDownloadButton"] button {
    background: #00e5a0 !important;
    color: #000000 !important;
    border: none !important;
    font-family: 'DM Mono', monospace !important;
    font-weight: 700 !important;
    font-size: 0.78rem !important;
    letter-spacing: 0.04em !important;
    border-radius: 6px !important;
    padding: 0 0.9rem !important;
    line-height: 1 !important;
    height: 2.4rem !important;
    min-height: 2.4rem !important;
    box-sizing: border-box !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
}
div.stButton > button p,
div.stButton > button[kind="primary"] p,
div.stButton > button[kind="primaryFormSubmit"] p,
[data-testid="stDownloadButton"] button p {
    color: #000000 !important;
    font-size: 0.78rem !important;
    font-family: 'DM Mono', monospace !important;
    font-weight: 700 !important;
    margin: 0 !important;
}
div.stButton > button:hover,
[data-testid="stDownloadButton"] button:hover { opacity: 0.85 !important; }

/* ── Red buttons: use type=secondary ── */
div.stButton > button[kind="secondary"],
div.stButton > button[kind="secondaryFormSubmit"] {
    background: #ff4d6d !important;
    color: #ffffff !important;
}
div.stButton > button[kind="secondary"] p,
div.stButton > button[kind="secondaryFormSubmit"] p {
    color: #ffffff !important;
}

/* ── Force buttons inside columns to fill their container ── */
div[data-testid="stColumn"] div.stButton,
div[data-testid="stColumn"] div.stButton > button,
div[data-testid="column"] div.stButton,
div[data-testid="column"] div.stButton > button {
    width: 100% !important;
}

/* ── Email anchor buttons — match Streamlit button height/positioning exactly ── */
a.email-btn {
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    background: #00e5a0 !important;
    color: #000 !important;
    font-family: 'DM Mono', monospace !important;
    font-size: 0.78rem !important;
    font-weight: 700 !important;
    padding: 0 0.9rem !important;
    border-radius: 6px !important;
    text-decoration: none !important;
    letter-spacing: 0.04em !important;
    box-sizing: border-box !important;
    width: 100% !important;
    height: 2.4rem !important;
    line-height: 1 !important;
    margin: 0 !important;
}
a.email-btn:hover { opacity: 0.85 !important; text-decoration: none !important; }
/* Zero out margins on the markdown container holding the email anchor so it aligns with adjacent button */
div[data-testid="stColumn"] [data-testid="stMarkdown"]:has(> div > a.email-btn),
div[data-testid="column"] [data-testid="stMarkdown"]:has(> div > a.email-btn) {
    margin: 0 !important;
    padding: 0 !important;
}

/* ── Ensure button rows align consistently — zero margins on element containers that hold buttons or email anchors ── */
div[data-testid="stColumn"] [data-testid="stElementContainer"]:has(div.stButton),
div[data-testid="stColumn"] [data-testid="stElementContainer"]:has([data-testid="stDownloadButton"]),
div[data-testid="stColumn"] [data-testid="stElementContainer"]:has(a.email-btn),
div[data-testid="column"] [data-testid="stElementContainer"]:has(div.stButton),
div[data-testid="column"] [data-testid="stElementContainer"]:has([data-testid="stDownloadButton"]),
div[data-testid="column"] [data-testid="stElementContainer"]:has(a.email-btn) {
    margin: 0 !important;
    padding: 0 !important;
}

/* File uploader - force the Browse/Upload button green */
.stFileUploader label { display: none !important; }
[data-testid="stFileUploaderDropzone"] button,
section[data-testid="stFileUploader"] button {
    background: #00e5a0 !important;
    color: #000000 !important;
    border: none !important;
    font-family: 'DM Mono', monospace !important;
    font-weight: 700 !important;
    letter-spacing: 0.04em !important;
    border-radius: 6px !important;
    padding: 0 0.9rem !important;
    line-height: 1 !important;
    height: 2.4rem !important;
    min-height: 2.4rem !important;
    position: relative !important;
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
}
/* Hide the original "Upload"/"Browse files" text but keep icons (svg) visible */
[data-testid="stFileUploaderDropzone"] button p,
[data-testid="stFileUploaderDropzone"] button span:not([data-testid]):not(:has(svg)),
section[data-testid="stFileUploader"] button p,
section[data-testid="stFileUploader"] button span:not([data-testid]):not(:has(svg)) {
    font-size: 0 !important;
    color: transparent !important;
    width: 0 !important;
    overflow: hidden !important;
}
/* Inject our own label — centered via the parent's flex layout */
[data-testid="stFileUploaderDropzone"] button::after,
section[data-testid="stFileUploader"] button::after {
    content: "Upload & Triage";
    font-size: 0.78rem !important;
    font-family: 'DM Mono', monospace !important;
    font-weight: 700 !important;
    color: #000000 !important;
    letter-spacing: 0.04em !important;
    display: block;
    text-align: center;
    width: 100%;
}
[data-testid="stFileUploaderDropzone"] button:hover,
section[data-testid="stFileUploader"] button:hover { opacity: 0.85 !important; }

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    background: transparent !important;
    border-bottom: 1px solid var(--border) !important;
    gap: 0 !important;
}

.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    color: var(--subtext) !important;
    font-family: var(--mono) !important;
    font-size: 0.78rem !important;
    letter-spacing: 0.06em !important;
    border-bottom: 2px solid transparent !important;
    padding: 0.75rem 1.5rem !important;
}

.stTabs [aria-selected="true"] {
    color: var(--accent) !important;
    border-bottom-color: var(--accent) !important;
}

.stTabs [data-baseweb="tab-panel"] {
    padding-top: 1.5rem !important;
}

/* Spinner */
.stSpinner > div { border-top-color: var(--accent) !important; }

/* Section label */
.section-label {
    font-family: var(--mono) !important;
    font-size: 0.65rem;
    letter-spacing: 0.15em;
    text-transform: uppercase;
    color: var(--text);
    margin-bottom: 0.75rem;
    padding-bottom: 0.5rem;
    border-bottom: 1px solid var(--border);
}

.stat-box {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 1rem 1.25rem;
    text-align: center;
}

.stat-num {
    font-family: var(--display) !important;
    font-size: 2rem;
    font-weight: 800;
    line-height: 1;
    margin-bottom: 0.25rem;
}

.stat-label {
    font-family: var(--mono) !important;
    font-size: 0.65rem;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: var(--subtext);
}
</style>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
VIP_USERS = {"c.osei"}

RUBRIC_QUESTIONS = [
    "Does the ticket explicitly name or describe more than one user or more than one site as being affected? Answer only YES or NO. If you are uncertain, answer YES.",
    "Does the ticket describe system behavior that contradicts what the system is supposed to do, such as showing wrong data, failing to send, or behaving differently after an update, affecting multiple users or the system as a whole — not just one user's browser or local device? Answer only YES or NO. If you are uncertain, answer YES.",
    "Does the ticket contain the words 'patient', 'clinical', or 'treatment' in a context that describes a current disruption? Answer only YES or NO. If you are uncertain, answer YES.",
    "Does the ticket explicitly request changes to slot capacity rules, approval workflows, or site-specific system configuration — not standard user account creation or access provisioning? Answer only YES or NO. If you are uncertain, answer YES.",
    "Does the ticket explicitly state that the same issue has happened before or is happening repeatedly? Answer only YES or NO. If you are uncertain, answer YES.",
]

RUBRIC_LABELS = [
    "Multiple users or sites",
    "System bug or data error",
    "Patient or clinical impact",
    "Custom config / workflow",
    "Recurring issue",
    "VIP user",
]

SYSTEM_CONTEXT = """You are an IT ticket triage assistant supporting a Senior Business
Systems Manager who oversees scheduling systems used by master schedulers at
manufacturing sites. These schedulers manage access to manufacturing slots for
made-to-order cell therapies — a time-critical, high-stakes operation where
system issues can directly impact patient treatment timelines."""

EMAIL_PROMPT = """You are drafting two very short, professional emails for an escalated IT support ticket.

Context about the ticket:
TICKET: {ticket}

Rubric findings:
{rubric_summary}

Draft two succinct emails:

EMAIL 1 - TO BUSINESS USER (from contracted IT team):
- Acknowledge receipt
- Confirm escalation to senior management
- Confirm IT is their point of contact
- One short paragraph max
- Subject line must start with exactly "ESCALATED - " followed by a brief description

EMAIL 2 - TO IT RESOURCE:
- Summarize the issue in one sentence
- Note senior manager oversight
- Flag any patient/clinical urgency if applicable
- Final line must say exactly: "Please provide a status update within the next two hours."
- Subject line must start with exactly "ESCALATED - " followed by a brief description

Respond in this exact format:

EMAIL 1 - TO BUSINESS USER:
Subject: ESCALATED - [brief description]
Body:
[email body]

EMAIL 2 - TO IT RESOURCE:
Subject: ESCALATED - [brief description]
Body:
[email body]
"""

# ── Session state ──────────────────────────────────────────────────────────────
if "escalation_log" not in st.session_state:
    st.session_state.escalation_log = []
if "results" not in st.session_state:
    st.session_state.results = []
if "processing_queue" not in st.session_state:
    st.session_state.processing_queue = []
if "is_processing" not in st.session_state:
    st.session_state.is_processing = False
if "action" not in st.session_state:
    st.session_state.action = None
if "action_idx" not in st.session_state:
    st.session_state.action_idx = None
if "last_processed_file" not in st.session_state:
    st.session_state.last_processed_file = None
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

# ── API helpers ────────────────────────────────────────────────────────────────
def get_client():
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        try:
            api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
        except Exception:
            api_key = ""
    if not api_key:
        st.error("ANTHROPIC_API_KEY not set. Add it to your environment or Streamlit secrets.")
        st.stop()
    return anthropic.Anthropic(api_key=api_key)

def ask_rubric_question(client, ticket, question):
    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=10,
        system=SYSTEM_CONTEXT,
        messages=[{"role": "user", "content": f"Ticket:\n{ticket}\n\nQuestion: {question}"}]
    )
    answer = message.content[0].text.strip().upper()
    if not answer.startswith("YES") and not answer.startswith("NO"):
        return "YES"
    return "YES" if answer.startswith("YES") else "NO"

def draft_emails(client, ticket, rubric_answers):
    rubric_summary = "\n".join(
        f"- {RUBRIC_LABELS[i]}: {rubric_answers[i]}"
        for i in range(len(rubric_answers))
    )
    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=600,
        system=SYSTEM_CONTEXT,
        messages=[{"role": "user", "content": EMAIL_PROMPT.format(
            ticket=ticket, rubric_summary=rubric_summary
        )}]
    )
    return message.content[0].text

DELEGATE_SUMMARY_PROMPT = """In 8 words or fewer, summarize the core issue in this IT support ticket. Be specific and concise. No punctuation at the end.

Ticket: {ticket}

Respond with only the summary, nothing else."""

def generate_summary(client, ticket):
    """Generate a short summary for delegate tickets."""
    message = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=30,
        system=SYSTEM_CONTEXT,
        messages=[{"role": "user", "content": DELEGATE_SUMMARY_PROMPT.format(ticket=ticket)}]
    )
    return message.content[0].text.strip()

def extract_subject(emails):
    """Extract subject from EMAIL 1, strip ESCALATED - prefix."""
    if not emails:
        return None
    for line in emails.split("\n"):
        if line.startswith("Subject:"):
            subject = line.replace("Subject:", "").strip()
            if subject.upper().startswith("ESCALATED -"):
                subject = subject[len("ESCALATED -"):].strip()
            elif subject.upper().startswith("ESCALATED-"):
                subject = subject[len("ESCALATED-"):].strip()
            return subject
    return None

def triage_ticket(client, ticket, username=""):
    rubric_answers = []
    for question in RUBRIC_QUESTIONS:
        answer = ask_rubric_question(client, ticket, question)
        rubric_answers.append(answer)
    is_vip = "YES" if username.lower() in VIP_USERS else "NO"
    rubric_answers.append(is_vip)
    classification = "ESCALATE" if "YES" in rubric_answers else "DELEGATE"
    emails = draft_emails(client, ticket, rubric_answers) if classification == "ESCALATE" else None
    if classification == "ESCALATE":
        summary = extract_subject(emails) or ticket[:60]
    else:
        summary = generate_summary(client, ticket)
    return classification, rubric_answers, emails, summary

def create_msg_file(subject, body):
    """Create a .msg-compatible file using RFC 2822 .eml format which opens in Outlook."""
    import io as _io
    timestamp = datetime.datetime.now().strftime("%a, %d %b %Y %H:%M:%S +0000")
    content = f"""MIME-Version: 1.0
Date: {timestamp}
Subject: {subject}
Content-Type: text/plain; charset="utf-8"

{body}
"""
    return content.encode("utf-8")

# ── Render helpers ─────────────────────────────────────────────────────────────
def render_rubric(rubric_answers):
    items = ""
    for i, label in enumerate(RUBRIC_LABELS):
        val = rubric_answers[i]
        val_html = f'<span class="rubric-yes">YES</span>' if val == "YES" else f'<span class="rubric-no">NO</span>'
        items += f'<div class="rubric-item"><span class="rubric-label">{label}</span>{val_html}</div>'
    st.markdown(f'<div class="rubric-grid">{items}</div>', unsafe_allow_html=True)

def render_result(result, index=None):
    ticket = result["ticket"]
    username = result.get("username", "")
    classification = result["classification"]
    rubric_answers = result["rubric_answers"]
    emails = result.get("emails", "")
    timestamp = result.get("timestamp", "")

    vip = " [VIP]" if username.lower() in VIP_USERS else ""
    user = f"@{username}{vip}  |  " if username else ""
    preview = ticket[:70] + "..." if len(ticket) > 70 else ticket
    label = f"{classification}  |  {user}{preview}"

    with st.expander(label):
        st.caption(f"Submitted: {timestamp}")
        st.text(ticket)
        render_rubric(rubric_answers)
        if classification == "ESCALATE" and emails:
            parts = emails.split("EMAIL 2 - TO IT RESOURCE:")
            email1 = parts[0].replace("EMAIL 1 - TO BUSINESS USER:", "").strip()
            email2 = parts[1].strip() if len(parts) > 1 else ""
            st.markdown("<div class=\"email-caption\">Email to Business User</div>", unsafe_allow_html=True)
            st.text_area("e1", email1, height=150, disabled=True, label_visibility="collapsed", key=f"res_email1_{id(result)}")
            if email2:
                st.markdown("<div class=\"email-caption\">Email to IT Resource</div>", unsafe_allow_html=True)
                st.text_area("e2", email2, height=150, disabled=True, label_visibility="collapsed", key=f"res_email2_{id(result)}")

def export_log_excel():
    import io as _io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Escalation Log"

    headers = ["Timestamp", "Username", "VIP", "Summary", "Ticket", "Classification"] + RUBRIC_LABELS + ["Email to Business User", "Email to IT Resource"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF4D6D")
        cell.alignment = Alignment(wrap_text=True)

    for entry in st.session_state.escalation_log:
        emails = entry.get("emails", "") or ""
        parts = emails.split("EMAIL 2 - TO IT RESOURCE:")
        email1 = parts[0].replace("EMAIL 1 - TO BUSINESS USER:", "").strip()
        email2 = parts[1].strip() if len(parts) > 1 else ""
        vip = "Yes" if entry.get("username", "").lower() in VIP_USERS else "No"
        row = [
            entry.get("timestamp", ""),
            entry.get("username", ""),
            vip,
            entry.get("summary", ""),
            entry.get("ticket", ""),
            entry.get("classification", ""),
        ] + [entry["rubric_answers"][j] for j in range(len(RUBRIC_LABELS))] + [email1, email2]
        ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 30

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def export_delegation_log_excel(delegated_all):
    import io as _io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delegation Log"

    headers = ["Timestamp", "Username", "Summary", "Ticket", "Classification"] + RUBRIC_LABELS
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="00A872")
        cell.alignment = Alignment(wrap_text=True)

    for entry in delegated_all:
        row = [
            entry.get("timestamp", ""),
            entry.get("username", ""),
            entry.get("summary", ""),
            entry.get("ticket", ""),
            entry.get("classification", ""),
        ] + [entry["rubric_answers"][j] for j in range(len(RUBRIC_LABELS))]
        ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 30

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="header-banner">
    <div class="header-title">Cell Therapy <span>Ticket Triage</span></div>
    <div class="header-sub">Global Supply Chain · Business Systems</div>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="disclaimer">
    ⚠ ESCALATE classifications are recommended for your review. Email drafts are suggestions only — review before sending. This system does not send emails automatically.
</div>
""", unsafe_allow_html=True)

col_rubric, col_input = st.columns([2, 1], gap="large")

with col_rubric:
    st.markdown("""
    <div style="margin-top: 0.25rem;">
        <div style="font-family: 'DM Mono', monospace; font-size: 0.72rem; font-weight:700; letter-spacing: 0.1em; text-transform: uppercase; color: #ffffff; border-bottom: 1px solid #1e2330; padding-bottom: 0.5rem; margin-bottom: 0.75rem;">
            Triage Rubric &mdash; <span style="color:#ff4d6d;">Any YES answer triggers ESCALATE</span>
        </div>
        <div style="font-family: 'DM Mono', monospace; font-size: 0.875rem; color: #ffffff; line-height: 1.7;">
            <div style="padding: 0.25rem 0; border-bottom: 1px solid #1e2330;"><span style="color:#00e5a0; font-weight:700;">Q1</span> &nbsp; Does the ticket explicitly name or describe more than one user or more than one site as being affected?</div>
            <div style="padding: 0.25rem 0; border-bottom: 1px solid #1e2330;"><span style="color:#00e5a0; font-weight:700;">Q2</span> &nbsp; Does the ticket describe system behavior that contradicts what the system is supposed to do, such as showing wrong data, failing to send, or behaving differently after an update, affecting multiple users or the system as a whole — not just one user's browser or local device?</div>
            <div style="padding: 0.25rem 0; border-bottom: 1px solid #1e2330;"><span style="color:#00e5a0; font-weight:700;">Q3</span> &nbsp; Does the ticket contain the words 'patient', 'clinical', or 'treatment' in a context that describes a current disruption?</div>
            <div style="padding: 0.25rem 0; border-bottom: 1px solid #1e2330;"><span style="color:#00e5a0; font-weight:700;">Q4</span> &nbsp; Does the ticket explicitly request changes to slot capacity rules, approval workflows, or site-specific system configuration — not standard user account creation or access provisioning?</div>
            <div style="padding: 0.25rem 0; border-bottom: 1px solid #1e2330;"><span style="color:#00e5a0; font-weight:700;">Q5</span> &nbsp; Does the ticket explicitly state that the same issue has happened before or is happening repeatedly?</div>
            <div style="padding: 0.25rem 0;"><span style="color:#00e5a0; font-weight:700;">Q6</span> &nbsp; Was this ticket submitted by a VIP user?</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

with col_input:
    tab_batch, tab_single = st.tabs(["Batch CSV Upload", "Single Ticket"])

    with tab_single:
        username = st.text_input("Username (optional)", placeholder="e.g. j.martinez", key="single_username", label_visibility="collapsed")
        ticket_text = st.text_area("Ticket", placeholder="Paste ticket here...", height=80, key="single_ticket", label_visibility="collapsed")
        if st.button("Triage →", key="single_submit", type="primary"):
            if not ticket_text.strip():
                st.warning("Please enter a ticket.")
            else:
                client = get_client()
                with st.spinner("Running rubric..."):
                    classification, rubric_answers, emails, summary = triage_ticket(client, ticket_text.strip(), username.strip())
                timestamp = datetime.datetime.now().strftime("%H:%M:%S")
                result = {
                    "ticket": ticket_text.strip(),
                    "username": username.strip(),
                    "classification": classification,
                    "rubric_answers": rubric_answers,
                    "emails": emails,
                    "summary": summary,
                    "timestamp": timestamp,
                }
                st.session_state.results.insert(0, result)
                if classification == "ESCALATE":
                    st.session_state.escalation_log.insert(0, result)

    with tab_batch:
        st.markdown("""
        <style>
        [data-testid="stFileUploaderDropzoneInstructions"] { display: none !important; }
        [data-testid="stFileUploaderDropzone"] {
            padding: 0 !important; min-height: 0 !important;
            border: none !important; background: transparent !important;
        }
        [data-testid="stFileUploaderDropzone"] small { display: none !important; }
        section[data-testid="stFileUploader"] { margin-bottom: 0 !important; }
        div[data-testid="stFileUploader"] > label { display: none !important; }
        </style>
        """, unsafe_allow_html=True)

        uploaded_file = st.file_uploader("Upload and Triage", type=["csv"], key=f"batch_upload_{st.session_state.uploader_key}", label_visibility="hidden")

        # Hide the upload button once a file is loaded — triage runs automatically
        if uploaded_file is not None:
            st.markdown("""
            <style>
            [data-testid="stFileUploaderDropzone"] button,
            section[data-testid="stFileUploader"] button { display: none !important; }
            </style>
            """, unsafe_allow_html=True)

        # Fixed-height container — always present so layout never shifts
        progress_placeholder = st.empty()
        status_placeholder = st.empty()
        st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)

        # Auto-triage as soon as a NEW file is uploaded (not on every rerun)
        file_signature = None
        if uploaded_file is not None:
            file_signature = f"{uploaded_file.name}:{uploaded_file.size}"

        if (
            uploaded_file
            and not st.session_state.is_processing
            and not st.session_state.processing_queue
            and st.session_state.get("last_processed_file") != file_signature
        ):
            content = uploaded_file.read().decode("utf-8")
            reader = csv.DictReader(io.StringIO(content))
            rows = [
                {"ticket": r.get("ticket", "").strip(), "username": r.get("username", "").strip()}
                for r in reader if r.get("ticket", "").strip()
            ]
            if "ticket" not in reader.fieldnames:
                st.error("CSV must contain a 'ticket' column.")
            else:
                st.session_state.processing_queue = rows
                st.session_state.is_processing = True
                st.session_state.last_processed_file = file_signature
                st.rerun()

        if st.session_state.is_processing and st.session_state.processing_queue:
            client = get_client()
            total = len(st.session_state.processing_queue)
            while st.session_state.processing_queue:
                done = total - len(st.session_state.processing_queue)
                progress_placeholder.progress(done / total)
                status_placeholder.caption(f"Processing ticket {done + 1} of {total}...")
                row = st.session_state.processing_queue[0]
                classification, rubric_answers, emails, summary = triage_ticket(client, row["ticket"], row["username"])
                timestamp = datetime.datetime.now().strftime("%H:%M:%S")
                result = {
                    "ticket": row["ticket"],
                    "username": row["username"],
                    "classification": classification,
                    "rubric_answers": rubric_answers,
                    "emails": emails,
                    "summary": summary,
                    "timestamp": timestamp,
                }
                st.session_state.results.insert(0, result)
                if classification == "ESCALATE":
                    st.session_state.escalation_log.insert(0, result)
                st.session_state.processing_queue.pop(0)
            st.session_state.is_processing = False
            progress_placeholder.empty()
            status_placeholder.caption("✓ All tickets processed.")
            st.session_state.uploader_key += 1
            st.rerun()


st.markdown("""
<script>
(function() {
    function fix() {
        // Red Remove and Clear All buttons (belt and suspenders alongside CSS)
        document.querySelectorAll('button').forEach(function(b) {
            var p = b.querySelector('p');
            var t = (p ? p.textContent : b.textContent).trim();
            if (t === 'Remove' || t === 'Clear All') {
                b.style.setProperty('background', '#ff4d6d', 'important');
                b.style.setProperty('color', '#ffffff', 'important');
                if (p) {
                    p.style.setProperty('color', '#ffffff', 'important');
                    p.style.setProperty('-webkit-text-fill-color', '#ffffff', 'important');
                }
            }
        });
    }
    fix();
    setInterval(fix, 150);
    new MutationObserver(fix).observe(document.body, {childList:true, subtree:true});
})();
</script>
""", unsafe_allow_html=True)

st.divider()

# ── Logs — full width ──────────────────────────────────────────────────────────
log_col1, log_col2 = st.columns(2, gap="large")


# ── Escalation log ─────────────────────────────────────────────────────────────
with log_col1:
    st.markdown('<div class="section-label">Escalation Log</div>', unsafe_allow_html=True)

    if not st.session_state.escalation_log:
        st.caption("No escalations this session.")
    else:
        escalation_count = len(st.session_state.escalation_log)
        st.markdown(f'<div class="stat-box" style="margin-bottom:1rem"><div class="stat-num" style="color:var(--danger)">{escalation_count}</div><div class="stat-label">Escalations this session</div></div>', unsafe_allow_html=True)

        to_remove_escalation = None
        for idx, entry in enumerate(st.session_state.escalation_log):
            vip = " [VIP]" if entry.get("username", "").lower() in VIP_USERS else ""
            user = f"@{entry['username']}{vip}  |  " if entry.get("username") else ""
            summary = entry.get("summary") or entry["ticket"][:60]
            with st.expander(f"{entry['timestamp']}  |  {user}{summary}"):
                st.text(entry["ticket"])
                render_rubric(entry["rubric_answers"])
                if entry.get("emails"):
                    parts = entry["emails"].split("EMAIL 2 - TO IT RESOURCE:")
                    email1 = parts[0].replace("EMAIL 1 - TO BUSINESS USER:", "").strip()
                    email2 = parts[1].strip() if len(parts) > 1 else ""

                    # Extract subject lines
                    subj1, body1 = "", email1
                    subj2, body2 = "", email2
                    for line in email1.split("\n"):
                        if line.startswith("Subject:"):
                            subj1 = line.replace("Subject:", "").strip()
                            body1 = email1[email1.find("Body:") + 5:].strip() if "Body:" in email1 else email1
                    for line in email2.split("\n"):
                        if line.startswith("Subject:"):
                            subj2 = line.replace("Subject:", "").strip()
                            body2 = email2[email2.find("Body:") + 5:].strip() if "Body:" in email2 else email2

                    st.markdown("<div class=\"email-caption\">Email to Business User</div>", unsafe_allow_html=True)
                    st.text_area("biz_user", email1, height=150, disabled=True, label_visibility="collapsed", key=f"esc_email1_{idx}")
                    if email2:
                        st.markdown("<div class=\"email-caption\">Email to IT Resource</div>", unsafe_allow_html=True)
                        st.text_area("it_res", email2, height=150, disabled=True, label_visibility="collapsed", key=f"esc_email2_{idx}")

                btn_col1, btn_col2, btn_col3 = st.columns(3, gap="small")
                with btn_col1:
                    if st.button("Remove", key=f"remove_esc_{idx}", use_container_width=True, type="secondary"):
                        to_remove_escalation = idx
                with btn_col2:
                    if entry.get("emails") and subj1:
                        import urllib.parse
                        mailto1 = f"mailto:?subject={urllib.parse.quote(subj1)}&body={urllib.parse.quote(body1)}"
                        st.markdown(f'<a class="email-btn" href="{mailto1}" target="_blank">📧 Email Business User</a>', unsafe_allow_html=True)
                with btn_col3:
                    if entry.get("emails") and subj2:
                        mailto2 = f"mailto:?subject={urllib.parse.quote(subj2)}&body={urllib.parse.quote(body2)}"
                        st.markdown(f'<a class="email-btn" href="{mailto2}" target="_blank">📧 Email IT Resource</a>', unsafe_allow_html=True)

        if to_remove_escalation is not None:
            entry = st.session_state.escalation_log[to_remove_escalation]
            st.session_state.escalation_log.pop(to_remove_escalation)
            st.session_state.results = [r for r in st.session_state.results if r is not entry]
            st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)
        col_exp, col_clr = st.columns(2, gap="small")
        with col_exp:
            st.download_button(
                label="Export",
                data=export_log_excel(),
                file_name=f"escalation_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="export_log",
                use_container_width=True
            )
        with col_clr:
            if st.button("Clear All", key="clear_log", use_container_width=True, type="secondary"):
                st.session_state.results = [r for r in st.session_state.results if r["classification"] != "ESCALATE"]
                st.session_state.escalation_log = []
                st.rerun()

    # ── Delegation log ─────────────────────────────────────────────────────────
with log_col2:
    st.markdown('<div class="section-label">Delegation Log</div>', unsafe_allow_html=True)

    delegated_all = [r for r in st.session_state.results if r["classification"] == "DELEGATE"]
    if not delegated_all:
        st.caption("No delegations this session.")
    else:
        st.markdown(f'<div class="stat-box" style="margin-bottom:1rem"><div class="stat-num" style="color:var(--accent)">{len(delegated_all)}</div><div class="stat-label">Delegations this session</div></div>', unsafe_allow_html=True)

        to_remove_delegation = None
        for idx, entry in enumerate(delegated_all):
            user = f"@{entry['username']}  |  " if entry.get("username") else ""
            summary = entry.get("summary") or entry["ticket"][:60]
            with st.expander(f"{entry.get('timestamp', '')}  |  {user}{summary}"):
                st.text(entry["ticket"])
                render_rubric(entry["rubric_answers"])
                if st.button("Remove", key=f"remove_del_{idx}", type="secondary", use_container_width=True):
                    to_remove_delegation = idx

        if to_remove_delegation is not None:
            delegation_entries = [r for r in st.session_state.results if r["classification"] == "DELEGATE"]
            entry_to_remove = delegation_entries[to_remove_delegation]
            st.session_state.results = [r for r in st.session_state.results if r is not entry_to_remove]
            st.session_state.escalation_log = [r for r in st.session_state.escalation_log if r is not entry_to_remove]
            st.rerun()

        st.markdown("<br>", unsafe_allow_html=True)
        col_exp2, col_clr2 = st.columns(2, gap="small")
        with col_exp2:
            st.download_button(
                label="Export",
                data=export_delegation_log_excel(delegated_all),
                file_name=f"delegation_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="export_delegation_log",
                use_container_width=True
            )
        with col_clr2:
            if st.button("Clear All", key="clear_delegation_log", use_container_width=True, type="secondary"):
                st.session_state.results = [r for r in st.session_state.results if r["classification"] != "DELEGATE"]
                st.rerun()
